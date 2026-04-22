"""文件处理工具核心模块"""

import json
import os
import re
import logging
import zipfile
import xml.etree.ElementTree as ET
from typing import Dict, List, Any, Optional

import fitz
import openpyxl

logger = logging.getLogger(__name__)


def sanitize_filename(filename: str) -> str:
    """安全化文件名，防止安全问题"""
    if not filename:
        return ""
    # 移除路径组件
    filename = os.path.basename(filename)
    # 替换潜在危险字符
    filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
    # 限制长度
    if len(filename) > 255:
        name, ext = os.path.splitext(filename)
        filename = name[:250] + ext
    return filename


def column_number_to_letter(col_num: int) -> str:
    """将数字列号转换为Excel字母列号（1->A, 2->B, 27->AA）"""
    result = ""
    while col_num > 0:
        col_num -= 1  # 调整为0-based
        result = chr(65 + col_num % 26) + result
        col_num //= 26
    return result


def parse_xlsx(file_path: str) -> Dict[str, Any]:
    """解析Excel文件，返回结构化数据

    Args:
        file_path: Excel文件路径

    Returns:
        dict: 包含表格数据和元信息的字典
    """
    try:
        logger.info(f"Parsing Excel file: {file_path}")

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active

        # 获取所有行数据
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel文件为空")

        # 处理表头
        headers = [str(cell) if cell is not None else f"Column_{i+1}" for i, cell in enumerate(rows[0])]

        # 处理数据行
        data_rows = []
        for i, row in enumerate(rows[1:], 1):
            row_dict = {"_row_index": i + 1}  # 1-based行号

            for j, cell_value in enumerate(row):
                if j < len(headers):
                    header = headers[j]
                    row_dict[header] = str(cell_value) if cell_value is not None else ""

            data_rows.append(row_dict)

        # 提取超链接
        hyperlinks = extract_hyperlinks(file_path)

        result = {
            "filename": os.path.basename(file_path),
            "headers": headers,
            "data": data_rows,
            "total_rows": len(data_rows),
            "hyperlinks": hyperlinks
        }

        logger.info(f"Successfully parsed Excel: {len(data_rows)} rows, {len(hyperlinks)} hyperlinks")
        return result

    except Exception as e:
        logger.error(f"Failed to parse Excel file {file_path}: {e}")
        raise


def extract_hyperlinks(file_path: str) -> Dict[str, str]:
    """从Excel文件中提取超链接信息

    Args:
        file_path: Excel文件路径

    Returns:
        dict: {单元格位置: 链接URL} 的映射
    """
    hyperlinks = {}

    try:
        # 使用zipfile读取Excel文件的内部结构
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            # 读取工作表关系文件
            try:
                rels_content = zip_file.read('xl/worksheets/_rels/sheet1.xml.rels')
                rels_root = ET.fromstring(rels_content)

                # 构建关系映射：ID -> URL
                relationships = {}
                for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                    rel_id = rel.get('Id')
                    target = rel.get('Target')
                    if rel_id and target and target.startswith('http'):
                        relationships[rel_id] = target

            except KeyError:
                # 如果没有关系文件，说明没有超链接
                logger.debug("No hyperlinks found in Excel file")
                return {}

            # 读取工作表文件
            try:
                sheet_content = zip_file.read('xl/worksheets/sheet1.xml')
                sheet_root = ET.fromstring(sheet_content)

                # 查找超链接定义
                for hyperlink in sheet_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}hyperlink'):
                    ref = hyperlink.get('ref')  # 单元格位置，如A1
                    rel_id = hyperlink.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')

                    if ref and rel_id and rel_id in relationships:
                        hyperlinks[ref] = relationships[rel_id]

            except KeyError:
                logger.debug("No worksheet found in Excel file")

    except Exception as e:
        logger.warning(f"Failed to extract hyperlinks from {file_path}: {e}")

    logger.debug(f"Extracted {len(hyperlinks)} hyperlinks from Excel")
    return hyperlinks


def get_hyperlink_for_cell(hyperlinks: Dict[str, str], row_index: int, col_header: str, headers: List[str]) -> Optional[str]:
    """获取指定单元格的超链接

    Args:
        hyperlinks: 超链接映射字典
        row_index: 行号（1-based）
        col_header: 列标题
        headers: 表头列表

    Returns:
        str or None: 超链接URL，如果没有则返回None
    """
    try:
        # 找到列的索引
        if col_header not in headers:
            return None

        col_index = headers.index(col_header) + 1  # 1-based

        # 转换为Excel单元格引用格式（如A1, B2）
        col_letter = column_number_to_letter(col_index)
        cell_ref = f"{col_letter}{row_index + 1}"  # +1因为Excel中第一行是表头

        return hyperlinks.get(cell_ref)

    except Exception as e:
        logger.warning(f"Failed to get hyperlink for cell {row_index},{col_header}: {e}")
        return None


def read_pdf_content(file_path: str) -> str:
    """读取PDF文件内容

    Args:
        file_path: PDF文件路径

    Returns:
        str: PDF文本内容
    """
    try:
        logger.debug(f"Reading PDF content from: {file_path}")

        doc = fitz.open(file_path)
        text_content = ""

        for page_num in range(doc.page_count):
            page = doc[page_num]
            text_content += page.get_text()

        doc.close()

        logger.debug(f"Successfully read PDF: {len(text_content)} characters")
        return text_content.strip()

    except Exception as e:
        logger.error(f"Failed to read PDF {file_path}: {e}")
        return ""


def save_results_to_excel(data_rows: List[Dict[str, Any]], output_path: str, headers: Optional[List[str]] = None) -> bool:
    """将结果数据保存为Excel文件

    Args:
        data_rows: 数据行列表
        output_path: 输出文件路径
        headers: 表头列表，如果为None则从第一行数据推导

    Returns:
        bool: 保存是否成功
    """
    try:
        if not data_rows:
            logger.warning("No data to save to Excel")
            return False

        # 创建新的工作簿
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 确定表头
        if headers is None:
            # 从第一行数据推导表头（排除内部字段）
            headers = [key for key in data_rows[0].keys() if not key.startswith('_')]

        # 写入表头
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)

        # 写入数据
        for row_idx, row_data in enumerate(data_rows, 2):  # 从第2行开始
            for col_idx, header in enumerate(headers, 1):
                value = _normalize_excel_cell_value(row_data.get(header, ""))
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # 保存文件
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        workbook.save(output_path)
        workbook.close()

        logger.info(f"Successfully saved {len(data_rows)} rows to Excel: {output_path}")
        return True

    except Exception as e:
        logger.error(f"Failed to save results to Excel {output_path}: {e}")
        return False


def _normalize_excel_cell_value(value: Any) -> Any:
    """将复杂结构转换为 Excel 可写入的标量值。"""
    if value is None:
        return ""
    if isinstance(value, dict):
        return str(value.get("link") or value.get("url") or value.get("text") or json.dumps(value, ensure_ascii=False))
    if isinstance(value, list):
        parts = [_normalize_excel_cell_value(item) for item in value]
        return "; ".join([str(part) for part in parts if str(part).strip()])
    return value


def create_backup_filename(original_filename: str, timestamp_suffix: str = None) -> str:
    """创建带时间戳的备份文件名

    Args:
        original_filename: 原始文件名
        timestamp_suffix: 时间戳后缀，如果为None则使用当前时间

    Returns:
        str: 备份文件名
    """
    try:
        import time

        if timestamp_suffix is None:
            timestamp_suffix = time.strftime("%Y%m%d_%H%M%S")

        name, ext = os.path.splitext(original_filename)
        backup_name = f"{name}_backup_{timestamp_suffix}{ext}"

        return sanitize_filename(backup_name)

    except Exception as e:
        logger.warning(f"Failed to create backup filename: {e}")
        return f"backup_{timestamp_suffix}.xlsx"


def validate_excel_headers(headers: List[str], required_headers: List[str]) -> List[str]:
    """验证Excel表头是否包含必需的列

    Args:
        headers: 实际的表头列表
        required_headers: 必需的表头列表

    Returns:
        list: 缺失的表头列表
    """
    missing_headers = []

    for required in required_headers:
        if required not in headers:
            missing_headers.append(required)

    if missing_headers:
        logger.warning(f"Missing required headers: {missing_headers}")

    return missing_headers


def get_file_size_mb(file_path: str) -> float:
    """获取文件大小（MB）

    Args:
        file_path: 文件路径

    Returns:
        float: 文件大小（MB）
    """
    try:
        size_bytes = os.path.getsize(file_path)
        size_mb = size_bytes / (1024 * 1024)
        return round(size_mb, 2)
    except Exception as e:
        logger.warning(f"Failed to get file size for {file_path}: {e}")
        return 0.0


def is_valid_file_type(filename: str, allowed_extensions: List[str] = None) -> bool:
    """检查文件类型是否有效

    Args:
        filename: 文件名
        allowed_extensions: 允许的文件扩展名列表，默认为['.xlsx', '.xls']

    Returns:
        bool: 文件类型是否有效
    """
    if allowed_extensions is None:
        allowed_extensions = ['.xlsx', '.xls']

    try:
        _, ext = os.path.splitext(filename.lower())
        return ext in [e.lower() for e in allowed_extensions]
    except Exception as e:
        logger.warning(f"Failed to validate file type for {filename}: {e}")
        return False
